from io import BytesIO
from uuid6 import uuid7
from openpyxl import Workbook, load_workbook

from src.application.task.interfaces.excel import ExcelProcessor
from src.domain.task.entities import Task

from src.domain.task.value_objects import (
    TaskID,
    Code,
    Name,
    Address,
    Indication,
    Implementer,
    Coordinates,
    Comment,
    Number,
)

from .utils import draw_report_header


class OpenpyxlProcessor(ExcelProcessor):
    def get_file_from_database(self, tasks: list) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        draw_report_header(worksheet)
        for task in tasks:
            worksheet.append(
                (
                    task.code, task.address, task.name, None,
                    task.number, task.previous_indication, task.current_indication,
                    task.completion_date, task.latitude, task.longitude,
                    task.near_photo_url, task.far_photo_url
                )
            )
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer

    def get_tasks_from_excel(self, workbook: bytes) -> list[Task]:
        workbook = load_workbook(workbook)
        worksheet = workbook.active
        return [
            Task.create_task(
                task_id=TaskID(uuid7()),
                code=Code(
                    str(worksheet.cell(row=row, column=1).value)
                    if worksheet.cell(row=row, column=1).value else None
                ),
                name=Name(
                    str(worksheet.cell(row=row, column=3).value)
                    if worksheet.cell(row=row, column=3).value else None
                ),
                address=Address(
                    str(worksheet.cell(row=row, column=2).value)
                    if worksheet.cell(row=row, column=5).value else None
                ),
                indication=Indication(
                    previous=worksheet.cell(row=row, column=6).value,
                    current=worksheet.cell(row=row, column=7).value
                ),
                implementer=Implementer(None),
                coordinates=Coordinates(
                    longitude=worksheet.cell(row=row, column=10).value,
                    latitude=worksheet.cell(row=row, column=9).value
                ),
                comment=Comment(None),
                number=Number(
                    str(worksheet.cell(row=row, column=5).value)
                    if worksheet.cell(row=row, column=5).value else None
                ),
            )
            for row in range(3, worksheet.max_row + 1)
        ]
