from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font


def draw_report_header(worksheet: Worksheet) -> None:
    draw_first_line_report_header(worksheet)
    draw_second_line_report_header(worksheet)
    set_size_report_header(worksheet)


def draw_first_line_report_header(worksheet: Worksheet) -> None:
    worksheet.merge_cells("A1:E1")
    worksheet.merge_cells("F1:G1")
    worksheet.merge_cells("H1:H2")
    worksheet.merge_cells("I1:J1")
    worksheet.merge_cells("K1:L1")

    worksheet.cell(row=1, column=1).value = "Информация об элементе сети"
    worksheet.cell(row=1, column=6).value = "Показания"
    worksheet.cell(row=1, column=8).value = "Дата обхода"
    worksheet.cell(row=1, column=9).value = "Координаты"
    worksheet.cell(row=1, column=11).value = "Фотографии"

    worksheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(row=1, column=6).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(row=1, column=8).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.cell(row=1, column=9).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(row=1, column=11).alignment = Alignment(horizontal="center", vertical="center")

    worksheet.cell(row=1, column=1).font = Font(bold=True)
    worksheet.cell(row=1, column=6).font = Font(bold=True)
    worksheet.cell(row=1, column=8).font = Font(bold=True)
    worksheet.cell(row=1, column=9).font = Font(bold=True)
    worksheet.cell(row=1, column=11).font = Font(bold=True)


def draw_second_line_report_header(worksheet: Worksheet) -> None:
    worksheet.cell(row=2, column=1).value = "Идентификационный код"
    worksheet.cell(row=2, column=2).value = "Адрес"
    worksheet.cell(row=2, column=3).value = "Наименование объекта сети"
    worksheet.cell(row=2, column=4).value = "Тип прибора учета"
    worksheet.cell(row=2, column=5).value = "Номер ПУ"
    worksheet.cell(row=2, column=6).value = "Предыдущие показания"
    worksheet.cell(row=2, column=7).value = "Текущие показания"
    worksheet.cell(row=2, column=9).value = "Долгота"
    worksheet.cell(row=2, column=10).value = "Широта"
    worksheet.cell(row=2, column=11).value = "Показания"
    worksheet.cell(row=2, column=12).value = "Счетчик"

    worksheet.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(row=2, column=2).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(row=2, column=3).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(row=2, column=4).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(row=2, column=5).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(row=2, column=6).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(row=2, column=7).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(row=2, column=9).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(row=2, column=10).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(row=2, column=11).alignment = Alignment(horizontal="center", vertical="center")
    worksheet.cell(row=2, column=12).alignment = Alignment(horizontal="center", vertical="center")

    worksheet.cell(row=2, column=1).font = Font(bold=True)
    worksheet.cell(row=2, column=2).font = Font(bold=True)
    worksheet.cell(row=2, column=3).font = Font(bold=True)
    worksheet.cell(row=2, column=4).font = Font(bold=True)
    worksheet.cell(row=2, column=5).font = Font(bold=True)
    worksheet.cell(row=2, column=6).font = Font(bold=True)
    worksheet.cell(row=2, column=7).font = Font(bold=True)
    worksheet.cell(row=2, column=9).font = Font(bold=True)
    worksheet.cell(row=2, column=10).font = Font(bold=True)
    worksheet.cell(row=2, column=11).font = Font(bold=True)
    worksheet.cell(row=2, column=12).font = Font(bold=True)


def set_size_report_header(worksheet: Worksheet) -> None:
    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 20
    worksheet.column_dimensions["C"].width = 30
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["E"].width = 20
    worksheet.column_dimensions["F"].width = 25
    worksheet.column_dimensions["G"].width = 25
    worksheet.column_dimensions["H"].width = 20
    worksheet.column_dimensions["I"].width = 20
    worksheet.column_dimensions["J"].width = 20
    worksheet.column_dimensions["K"].width = 20
    worksheet.column_dimensions["L"].width = 20

    worksheet.row_dimensions[1].height = 30
    worksheet.row_dimensions[2].height = 30
